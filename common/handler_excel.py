
import openpyxl



# 创建一个操作excel的类
class HandleXlsx:

    def __init__(self, filename, sheetname):
        """
        操作xlsx文件的封装
        :param filename: 文件名或者路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读取xlsx文件的方法
        :return: 返回读取的内容并返回数据格式为列表嵌套字典的形式
        """
        # 加载工作簿对象
        wookbook = openpyxl.load_workbook(self.filename)
        sh = wookbook[self.sheetname]
        # 读取所有数据，转换为列表
        res = list(sh.rows)
        # 获取第一行标题
        title = [i.value for i in res[0]]
        cases = []
        # 遍历除第一行之外的数据
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        # 返回读取的数据
        return cases

    def write_data(self, row, column, value):
        """
        写入xlsx文件的方法
        :param row: 写入的行
        :param column:写入的列
        :param value: 写入的内容
        """
        # 加载工作簿对象
        wookbook = openpyxl.load_workbook(self.filename)
        sh = wookbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        wookbook.save(self.filename)


